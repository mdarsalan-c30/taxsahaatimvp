#!/usr/bin/env python3
"""Build the UX Designer Research Pack: a CSV bundle + a multi-tab xlsx.

Existing-doc rows are embedded below (extracted from docs/* and sibling
itr-filing-wireframes/*). External forum research is read from
docs/ux-research/_forum_research_raw.json (produced by the research agents).

Run from the lastminute-itr repo root:
    python3 scripts/build_ux_research_pack.py

Outputs:
    docs/ux-research/csv/*.csv        (11 sheets)
    docs/ux-research/UX_DESIGNER_RESEARCH_PACK_2026.xlsx
"""
import csv
import json
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:  # pragma: no cover - documented fallback
    HAS_OPENPYXL = False

OUT_DIR = os.path.join("docs", "ux-research")
CSV_DIR = os.path.join(OUT_DIR, "csv")
FORUM_JSON = os.path.join(OUT_DIR, "_forum_research_raw.json")
XLSX_PATH = os.path.join(OUT_DIR, "UX_DESIGNER_RESEARCH_PACK_2026.xlsx")

# Each tab: (sheet_name, [header...], [ [cell,...], ... ]). Populated below.
TABS = []

# ---------------------------------------------------------------------------
# Tab 0 - README (orientation for the designer)
# ---------------------------------------------------------------------------
TABS.append(("README", ["Section", "Detail"], [
    ["Product one-liner", "LastMinute ITR is an ITR companion, not an e-filer. We reconcile Form 16 / AIS / 26AS, compare old vs new regime, and guide the taxpayer screen-by-screen on incometax.gov.in. The user always files and e-verifies themselves."],
    ["Compliance copy rules", "No 'we file for you', no 'guaranteed refund', no ITD affiliation. Source: lib/copy/companion.ts."],
    ["How to read this pack", "11 tabs. Each tab has consistent, filterable columns. One idea per row. Severity = P0/P1/P2; Frequency/Relevance = High/Med/Low."],
    ["Column legend - Severity", "P0 = core to wedge / wrong-filing risk; P1 = high-value clarity; P2 = nice-to-have / SEO."],
    ["Column legend - Status", "Strong = well covered; Partial = exists but thin; Not_built = gap; Orphan/Redirect = dead route."],
    ["WHAT EXISTS", "Competitor analysis (rich), secondary user pain synthesis from qna.tax, full IA + funnel audit, 8 personas, strong design system, competitor filing-journal teardowns."],
    ["GAP - no interviews", "No primary user interviews conducted. Only a planned '30 users + 5 CAs' note. See Interviews tab + recommended script."],
    ["GAP - no survey data", "Feedback form is in-product only (1-5 stars, 280 chars). No analyzed survey/response dataset exists yet. See Surveys_Feedback tab."],
    ["GAP - competitor auth wall", "Quicko/ClearTax filing journals stop at the login/OTP wall. In-app pixel detail not captured. Forum quotes are the best proxy for real user voice."],
    ["Must-read sources", "docs/DESIGN_SYSTEM.md, itr-filing-wireframes/USER_FLOWS.md, docs/research/USER_PAINPOINTS_2026.md, docs/UX_IMPROVEMENT_PLAN.md, docs/FUNNEL_AUDIT_AND_SIMPLIFICATION.md."],
    ["Wireframe prototype", "itr-filing-wireframes/prototype/index.html (?flow=happy | ?flow=mismatch | ?viewport=mobile|desktop)."],
    ["MVP build", "Live MVP at taxsahaatimvp-exkv.vercel.app. Routes live under /file/*."],
]))

# ---------------------------------------------------------------------------
# Tab 1 - Competitors
# ---------------------------------------------------------------------------
TABS.append(("Competitors", [
    "Name", "Type", "URL", "Positioning", "Pricing_signal", "Strengths",
    "Weaknesses", "Our_wedge", "Audit_depth", "Source_file",
], [
    ["incometax.gov.in (DIY)", "India_direct", "https://www.incometax.gov.in", "Official free e-filing + e-verify", "Free", "Authoritative; pre-fill; the only place you can actually file", "Session timeouts ('Please Wait' loops), confusing utilities, zero hand-holding", "Calm copy-ready walkthrough that mirrors the portal's exact section order", "Landscape_only", "research/SIMILAR_PLATFORMS.md"],
    ["ClearTax (Clear)", "India_direct", "https://cleartax.in", "DIY + autofill, broad; 'File ITR in 3 simple steps'", "Rs 349-4000+ tiered", "Form 16 visual hero; multi Form 16 merge; capital gains/crypto; trust volume", "Pricier; tier upsell maze; pay before mismatch shown; discontinuing TaxCloud 2026", "Mismatch-first reconcile before any copy to portal; value-before-pay", "Full_journal", "COMPETITOR_ANALYSIS.md; competitor-audit/CLEARTAX_FILING_JOURNAL.md"],
    ["Quicko", "India_direct", "https://quicko.com", "DIY + broker autofill, trader-focused; 'File in ~15 min'", "Free salaried; ~Rs 799-999 paid; expert Rs 2499", "Broker import (Zerodha/Angel One); F&O turnover wizard; ITD autofill; expandable income nav", "Smaller broker list than ClearTax; slower July support; mismatch is a step not the moat", "Mismatch as hero + honest companion (not auto-file)", "Full_journal", "COMPETITOR_ANALYSIS.md; competitor-audit/QUICKO_FILING_JOURNAL.md"],
    ["TaxBuddy", "India_direct", "https://taxbuddy.in", "Software + assigned CA review", "Rs 500-2500", "AI pre-fill + human CA hand-off", "CA quality varies; turnaround slows in July; pricing confusion", "Self-serve clarity without forced CA dependency", "Landscape_only", "research/SIMILAR_PLATFORMS.md"],
    ["Tax2Win", "India_direct", "https://tax2win.in", "DIY free + add-ons; CA-assisted", "Free + add-ons", "One-click Form 16; fast filing", "Limited free features; add-ons required for complexity", "Honest single companion unlock, no add-on surprises", "Landscape_only", "research/SIMILAR_PLATFORMS.md"],
    ["TaxSpanner / myITreturn", "India_direct", "https://myitreturn.com", "Established DIY + assisted", "Rs 399-2999", "Reliable; assisted options", "Dated UX; no standout differentiator", "Premium calm UX + reconcile wedge", "Landscape_only", "research/SIMILAR_PLATFORMS.md"],
    ["EZTax", "India_direct", "https://eztax.in", "DIY + assisted, value pricing", "Low-cost tiers", "Cheap; broad coverage", "Cluttered UX; trust signals weak", "Trust-at-anxiety design; clarity", "None", "GAP_MATRIX_2026.md"],
    ["KoinX / TaxNodes", "India_adjacent", "https://koinx.com", "Crypto-specialist P&L", "Varies", "Crypto P&L computation", "Narrow scope; not full ITR", "Full ITR companion with CG as first-class head", "None", "research/SIMILAR_PLATFORMS.md"],
    ["qna.tax (Quicko forum)", "India_adjacent", "https://qna.tax", "Community Q&A (Discourse)", "Free", "Real answers to real edge cases; searchable", "Not a filing tool; answers scattered across threads", "Turn recurring threads into in-product help + guided flow", "Landscape_only", "research/USER_PAINPOINTS_2026.md"],
    ["TaxSaathi / TaxSathi", "India_adjacent", "https://taxsaathi.com", "Name-collision check (brand)", "n/a", "n/a - brand collision reference", "Naming overlap risk for our brand", "Distinct companion positioning + name decision", "None", "brand/BRAND_NAMING_2026.md"],
    ["TurboTax (Intuit)", "Global_UX_ref", "https://turbotax.intuit.com", "Guided interview wizard (US)", "Tiered SKUs", "Interview gold standard; one Q per screen; filing-ready % before pay", "US-only; late paywall; refund-max hype", "India AIS/26AS reconcile; pay after value (we reject refund hype)", "Landscape_only", "COMPETITOR_ANALYSIS.md; itr-filing-wireframes/COMPETITOR_GLOBAL.md"],
    ["H&R Block", "Global_UX_ref", "https://www.hrblock.com", "DIY + expert + in-store (US)", "SKU ladder", "Persistent help pane; income-first wizard; mode switching", "US forms; expert-first tone; max-refund hype", "Salaried DIY wedge; honest optimization", "Landscape_only", "COMPETITOR_ANALYSIS.md"],
    ["Stripe", "Global_UX_ref", "https://stripe.com", "Payments infra; activation = first charge", "n/a", "Incremental onboarding; inline validation; dashboard density; activation KPI", "Not tax; dev-oriented", "Treat Form 16 upload like 'first charge'; defer PAN/OTP", "Landscape_only", "COMPETITOR_ANALYSIS.md; DESIGN_SYSTEM.md"],
    ["Mercury", "Global_UX_ref", "https://mercury.com", "Startup banking; two-phase onboarding", "n/a", "Two-phase gate; real-time doc parse feedback; demo before signup; no upsell", "Not tax; US", "Phase 1 = name + Form 16; Phase 2 = PAN/bank at pay", "Landscape_only", "COMPETITOR_ANALYSIS.md; DESIGN_SYSTEM.md"],
    ["Ramp", "Global_UX_ref", "https://ramp.com", "Corporate finance automation", "n/a", "Contextual trust placement; AI reasoning transparency; bento marketing; itemized proof", "Not tax; SMB finance", "Trust cards at upload/mismatch/pay; AI autonomy guardrails", "Landscape_only", "COMPETITOR_ANALYSIS.md; DESIGN_SYSTEM.md"],
    ["Linear", "Global_UX_ref", "https://linear.app", "Product storytelling; motion restraint", "n/a", "Density + purposeful motion; clean light/dark", "Not fintech", "Section rhythm; status affordances; motion communicates state", "None", "DESIGN_SYSTEM.md"],
]))

# ---------------------------------------------------------------------------
# Tab 2 - User_Pain_Points (9 themes from qna.tax + landscape synthesis)
# ---------------------------------------------------------------------------
TABS.append(("User_Pain_Points", [
    "Theme_ID", "Theme", "User_quote_or_paraphrase", "Why_it_hurts", "Frequency",
    "Severity", "Our_route_today", "Gap", "Phase_2_UI_action", "Source",
], [
    ["T1", "AIS vs Form 16 vs 26AS reconciliation", "Form 16 salary is correct but AIS shows extra interest/dividend - what do I do? Will a mismatch trigger 143(1) or delay my refund?", "CPC auto-compares filed ITR vs AIS; mismatch triggers 143(1)(a) intimation and refund delay. Most DIY filers do not know AIS exists.", "High", "P0", "/file/import/mismatch, lib/filing/reconciliation.ts, /file/review taxes tab", "Three-doc model not shown side-by-side anywhere as one reconcile view", "/file/review reconcile dashboard: Form 16 vs AIS vs 26AS with matched/needs-attention states; let users file with correct figures + log AIS feedback", "research/USER_PAINPOINTS_2026.md"],
    ["T2", "Capital gains, intraday & F&O", "How are intraday and F&O profits taxed? F&O turnover and audit threshold? STCG/LTCG discrepancy and rebate under new regime?", "Form selection (ITR-2 vs ITR-3), turnover math, and 111A/112A special rates are genuinely hard; where users most fear notices.", "High", "P1", "Quiz/profiler routes ITR-2/3; /file/regime; /file/cabrain; engine handles 111A/112A", "CG data entry is estimate-placeholder only", "Review screen states 'capital gains detected -> ITR-2/3, here is what the portal needs'; honest 'complex -> consider expert'; CG as first-class income head", "research/USER_PAINPOINTS_2026.md"],
    ["T3", "Freelancer / business income (44ADA, foreign clients)", "Can I claim home-office expenses? Freelancer with foreign clients - how to report, is GST needed? High income low tax under 44ADA?", "Presumptive eligibility, 50% deemed profit, advance tax, GST thresholds confusing; foreign receipts add FEMA/GST nuance.", "Med", "P1", "Profiler routes presumptive (business code w); /file/other, /file/cabrain", "Path exists but thin on guidance", "Clear presumptive explainer + honest 'foreign income/GST -> consider expert'; covered mostly via learn/help", "research/USER_PAINPOINTS_2026.md"],
    ["T4", "New regime rebate interaction (FY 2025-26)", "New Regime FY25-26: calculation discrepancy on STCG/LTCG tax & 87A rebate eligibility. Old vs new - which is better for me?", "87A rebate does not apply to special-rate income (111A/112A); the interaction surprises people. Regime choice is highest-leverage for salaried.", "High", "P0", "/file/regime compare; RegimeCompareCard on landing; engine computes both regimes", "Strong already - keep", "Reconcile dashboard surfaces regime recommendation inline with honest estimate disclaimer and the 'why' (deductions lost in new, breakeven)", "research/USER_PAINPOINTS_2026.md"],
    ["T5", "Gifts, HUF, and 'do I report this?'", "Is a gift > Rs 50,000 from wife's parents tax-free in my HUF? Do I report gifts from parents? How are gifted shares/gold/property taxed?", "Relative-vs-non-relative exemption, HUF treatment, reporting obligations are murky; under-reporting risk is real.", "Med", "P2", "/file/other; glossary; learn articles", "Mostly a content/education gap", "Covered via Phase 1 learn/help + glossary; no Phase 2 UI blocker", "research/USER_PAINPOINTS_2026.md"],
    ["T6", "NRI & property TDS", "TDS on property purchased from an NRI seller (195 vs 194-IA). NRI reporting of Indian let-out property and interest.", "Higher TDS rates, Form 27Q, residential-status rules. High stakes, low confidence.", "Med", "P2", "Profiler supports non_resident -> ITR-2; /file/house-property", "Supported but advanced", "Honest escalation messaging; content coverage; no Phase 2 blocker", "research/USER_PAINPOINTS_2026.md"],
    ["T7", "The official portal itself is painful", "Stuck on 'Please Wait' loop / session timeout downloading Form 16C on TRACES. After half a day on the ITD site I used a tool and connected my broker - done.", "Strongest validation of companion mode: users do not hate filing, they hate portal friction and not knowing what to enter where.", "High", "P0", "/file/companion Portal Footprint Wizard; lib/engine/portalSections.ts; copy-ready field values", "This is our moat - keep & deepen", "Guided 'Path' bar + review->companion hand-off: always show 'you are N steps from filing on the portal'", "research/USER_PAINPOINTS_2026.md"],
    ["T8", "Pricing transparency & trust", "Users reward 'upload Form 16 + connect broker -> done'; punish add-on cost surprises and slow July support. Official portal is free; paid tools must justify the fee.", "Trust debt when value is not shown before pay; tier fatigue (ClearTax).", "Med", "P1", "/file/checkout/plans, value-before-pay gate (pay after review), PricingSection", "Value-before-pay model is correct", "Keep pricing legible and honest (companion walkthrough unlock, not government submission); reinforce in trust polish", "research/USER_PAINPOINTS_2026.md"],
    ["T9", "Which ITR form do I file?", "'Know your ITR' is one of Quicko's most-used tools; choosing ITR-1/2/3/4 is a recurring blocker.", "Wrong form = defective return / notice. Recurring entry-point anxiety.", "High", "P1", "ITR-type quiz on landing + /tools; profiler in onboarding", "Covered", "Ensure quiz and profiler outcomes are visible and reassuring in the funnel", "research/USER_PAINPOINTS_2026.md"],
    ["T-extra", "Multi-employer aggregation (job change)", "Two Form 16s after a job change - combined TDS is short because both employers gave the basic exemption separately, creating a payable surprise.", "Users expect a refund and get a demand; trust shock if not surfaced early.", "Med", "P1", "lib/filing/employers.ts; multi Form 16 upload; review 'Combined across N employers'", "Aggregation exists; surprise not explained early enough", "Per-employer breakdown + explain the payable surprise on review", "case-studies/CASE_STUDIES_2026.md (CS-2)"],
]))

# ---------------------------------------------------------------------------
# Tab 4 - Personas (CS-1..CS-8 + fixtures + marketing carousel)
# ---------------------------------------------------------------------------
TABS.append(("Personas", [
    "ID", "Name", "Archetype", "ITR_form", "Key_documents", "Pain_theme",
    "Expected_journey_A-F", "Screen_must_show", "Fixture_ID", "Source",
], [
    ["CS-1", "Aditi", "Single-employer salaried (simplest)", "ITR-1", "One Form 16", "T4 regime, T9 form", "A-F clean happy path", "Regime recommendation + 'you are filing ITR-1' confidence + portal walkthrough", "PersonaForm16 (base)", "case-studies/CASE_STUDIES_2026.md; e2e/fixtures/personas.ts"],
    ["CS-2", "Rohan", "Job change, two Form 16s", "ITR-1/2", "Two Form 16s", "T-extra multi-employer", "A-F with aggregation gate", "Per-employer TDS breakdown + 'why you now owe tax' explanation before pay", "PersonaForm16 x2", "case-studies/CASE_STUDIES_2026.md (CS-2)"],
    ["CS-3", "Meera", "AIS mismatch (extra interest/dividend)", "ITR-1", "Form 16 + AIS", "T1 reconciliation", "A-F, hard stop at mismatch", "Form 16 vs AIS side-by-side; resolve before copy to portal", "PersonaAisLine", "case-studies/CASE_STUDIES_2026.md (CS-3)"],
    ["CS-4", "Vikram", "Investor - capital gains (equity/MF)", "ITR-2", "Form 16 + broker P&L + AIS", "T2 capital gains", "A-F, ITR-2 route", "'CG detected -> ITR-2'; 111A/112A rates; what the portal CG schedule needs", "PersonaDocuments (CG)", "case-studies/CASE_STUDIES_2026.md (CS-4)"],
    ["CS-5", "Sana", "F&O / intraday trader", "ITR-3", "Broker P&L + turnover + AIS", "T2 F&O", "A-F, ITR-3 route", "Turnover/audit threshold honesty + 'complex -> consider expert' option", "PersonaDocuments (F&O)", "case-studies/CASE_STUDIES_2026.md (CS-5)"],
    ["CS-6", "Karan", "Freelancer / 44ADA presumptive", "ITR-4", "Bank + invoices + AIS", "T3 freelancer", "A-F, presumptive route", "44ADA 50% deemed profit explainer + advance tax note", "PersonaDocuments (business)", "case-studies/CASE_STUDIES_2026.md (CS-6)"],
    ["CS-7", "Priya", "House property / home loan", "ITR-1/2", "Form 16 + interest cert + AIS", "T6 property", "A-F with HP section", "Let-out vs self-occupied; 24(b) interest cap; portal HP schedule guide", "PersonaDocuments (HP)", "case-studies/CASE_STUDIES_2026.md (CS-7)"],
    ["CS-8", "Arjun", "NRI with Indian income", "ITR-2", "AIS + property/interest + TRC", "T6 NRI", "A-F, non-resident route", "Residential status gate + honest escalation for complex NRI cases", "PersonaFixture (non_resident)", "case-studies/CASE_STUDIES_2026.md (CS-8)"],
    ["MC-1", "Carousel: Salaried", "Marketing persona - salaried", "ITR-1", "Form 16", "T4/T9", "Landing -> A", "RegimeCompareCard; '15-minute' honest framing", "n/a", "lib/copy/competitorInspired.ts (PERSONA_CAROUSEL)"],
    ["MC-2", "Carousel: Trader", "Marketing persona - trader/investor", "ITR-2/3", "Broker import", "T2", "Landing -> A", "Import strip (broker connect); CG/F&O reassurance", "n/a", "lib/copy/competitorInspired.ts"],
    ["MC-3", "Carousel: Freelancer", "Marketing persona - freelancer", "ITR-4", "Bank + invoices", "T3", "Landing -> A", "Presumptive explainer entry; honest scope", "n/a", "lib/copy/competitorInspired.ts"],
]))

# ---------------------------------------------------------------------------
# Tab 5 - Surveys_Feedback
# ---------------------------------------------------------------------------
TABS.append(("Surveys_Feedback", [
    "Instrument", "Type", "Location_in_app", "Fields", "Validation_rules",
    "API_route", "Public_reviews_page", "Data_available", "Notes",
], [
    ["FeedbackScreen", "in_app", "Marketing / post-flow CTA -> FeedbackScreen", "rating (1-5 stars), message (textarea)", "message max 280 chars; rating required to enable submit", "POST /api/feedback", "n/a", "No", "Submissions go to /api/feedback; no analyzed/exported dataset exists yet. Source: components/marketing/FeedbackScreen.tsx"],
    ["companion_footprint_step_viewed", "in_app (analytics)", "Portal Footprint Wizard steps", "step_id, step_index, total_steps, persona", "event fires on step view", "PostHog capture", "n/a", "Partial (event defined)", "Tells which portal steps users dwell on / drop. Source: docs/MCKINSEY_ENGAGEMENT/06_ANALYTICS_EVENTS.md"],
    ["companion_field_action", "in_app (analytics)", "Companion field rows", "field_id, action (copy/expand/help)", "fires on field interaction", "PostHog capture", "n/a", "Partial (event defined)", "Signals which copy-ready fields are used most."],
    ["companion_field_copy", "in_app (analytics)", "Copy-to-clipboard buttons", "field_id, value_type", "fires on copy", "PostHog capture", "n/a", "Partial (event defined)", "Direct proxy for portal walkthrough value delivered."],
    ["companion_field_confusion", "in_app (analytics)", "Field help / confusion control", "field_id, reason", "fires on confusion tap", "PostHog capture", "n/a", "Partial (event defined)", "Flags fields needing clearer copy - feed straight into redesign."],
    ["User interview survey", "none", "Not built", "n/a", "n/a", "n/a", "n/a", "No", "GAP: no structured survey instrument and no interview responses exist. Designer should build one (see Interviews tab)."],
    ["Public reviews / testimonials", "planned", "Landing testimonials section (copy only)", "n/a", "n/a", "n/a", "Landing testimonials (static copy, not user-submitted)", "No", "Testimonials are marketing copy, not a live review dataset."],
]))

# ---------------------------------------------------------------------------
# Tab 6 - Interviews (none conducted; recommended script for the designer)
# ---------------------------------------------------------------------------
TABS.append(("Interviews", [
    "Status", "Planned_count", "Conducted_count", "Participant_type",
    "Script_outline", "Key_questions", "Output_location",
], [
    ["Not conducted", "30 users + 5 CAs (planned only)", "0", "Salaried DIY filers", "Intro -> last-year filing recall -> companion concept reaction -> mismatch scenario -> pricing reaction -> wrap", "Q1: Walk me through how you filed last year. Q2: What was the most confusing moment? Q3: Did you know AIS/26AS existed before filing?", "TBD - designer to create docs/ux-research/interviews/"],
    ["Not conducted", "-", "0", "Investors / traders (CG, F&O)", "Same spine, probe form-selection anxiety and notice fear", "Q4: How confident were you picking ITR-1/2/3/4? Q5: Have you ever received a 143(1) intimation - how did it feel?", "TBD"],
    ["Not conducted", "-", "0", "Freelancers (44ADA)", "Same spine, probe presumptive + advance tax confusion", "Q6: How do you decide what income to report? Q7: Did you understand presumptive taxation?", "TBD"],
    ["Not conducted", "-", "0", "CAs / tax practitioners", "Probe where clients get stuck; what a companion should never claim", "Q8: Where do DIY clients most often go wrong? Q9: What would make you trust/distrust a companion tool?", "TBD"],
    ["Recommended script", "-", "-", "All", "Companion comprehension test", "Q10: After seeing this, who actually files your return - you or the tool? (tests companion comprehension). Q11: At what point would you expect to pay? (paywall timing). Q12: Was the portal step-by-step guide useful or redundant?", "Use as moderated test once a prototype exists"],
]))

# ---------------------------------------------------------------------------
# Tab 7 - Information_Architecture
# ---------------------------------------------------------------------------
TABS.append(("Information_Architecture", [
    "Zone", "Route", "Screen_name", "Indexed", "Journey_step_A-F",
    "Primary_CTA", "Nav_parent", "Status", "Notes",
], [
    ["Marketing", "/", "Landing / Home", "Yes", "Pre-A", "Start / Upload Form 16", "Top nav", "Live", "Hero, RegimeCompareCard, persona carousel, pricing, testimonials"],
    ["Marketing", "/learn", "Learn index", "Yes", "Pre-A", "Read article", "Top nav", "Live", "SEO blog hub (LearnArticle objects); mirrored at /blogs"],
    ["Marketing", "/learn/[slug]", "Learn article", "Yes", "Pre-A", "Start filing", "Learn", "Live", "Content -> filing funnel entry"],
    ["Marketing", "/blogs/[slug]", "Blog article (mirror)", "Yes", "Pre-A", "Start filing", "Learn", "Live", "Mirror route of /learn/[slug]"],
    ["Marketing", "/tools", "Tools hub", "Yes", "Pre-A", "Use tool", "Top nav", "Live", "ITR-type quiz, regime calculator entry points"],
    ["Marketing", "/help", "Help center", "Yes", "Pre-A", "Search help", "Footer", "Live", "Help articles; sitemap-driven"],
    ["Marketing", "/glossary", "Glossary", "Yes", "Pre-A", "-", "Footer", "Live", "Term definitions for SEO + in-product linking"],
    ["Filing", "/file/onboarding", "Profiler / onboarding", "No", "A", "Continue", "Funnel", "Live", "Profiler routes ITR-1/2/3/4; persona detection"],
    ["Filing", "/file/import", "Import / upload", "No", "B", "Upload Form 16 / connect", "Funnel", "Live", "Form 16 upload + connector strip (import value = 'first charge')"],
    ["Filing", "/file/import/mismatch", "Mismatch resolve", "No", "B", "Resolve", "Funnel", "Live", "AIS vs Form 16 reconciliation - the wedge"],
    ["Filing", "/file/regime", "Regime compare", "No", "C", "Choose regime", "Funnel", "Live", "Old vs new computed both ways"],
    ["Filing", "/file/review", "Review / reconcile dashboard", "No", "D", "Continue to pay", "Funnel", "Live", "Three-doc reconcile; taxes tab; regime inline"],
    ["Filing", "/file/checkout/plans", "Plans / pay", "No", "E", "Pay", "Funnel", "Live", "Value-before-pay gate (pay after review)"],
    ["Filing", "/file/companion", "Portal Footprint Wizard", "No", "F", "Copy to portal", "Funnel", "Live", "Screen-by-screen incometax.gov.in guide; copy-ready fields - the moat"],
    ["Filing", "/file/cabrain", "CA Brain", "No", "C/D", "-", "Funnel", "Orphan/Review", "Funnel audit flags as candidate to merge/cut; advanced reasoning surface"],
    ["Filing", "/file/everify", "E-verify guide", "No", "Post-F", "E-verify on portal", "Funnel", "Redirect/Thin", "Funnel audit flags as dead/thin route; user e-verifies on portal"],
    ["Filing", "/file/house-property", "House property", "No", "C", "Continue", "Funnel", "Live", "HP schedule support (CS-7)"],
    ["Filing", "/file/other", "Other income / disclosures", "No", "C", "Continue", "Funnel", "Live", "Gifts, other sources, misc heads"],
    ["API", "/api/feedback", "Feedback endpoint", "No", "n/a", "-", "n/a", "Live", "Receives FeedbackScreen submissions"],
    ["API", "/sitemap.xml", "Sitemap", "Yes", "n/a", "-", "n/a", "Live", "Static + dynamic (articles/blogs/glossary/help). Source: app/sitemap.ts"],
]))

# ---------------------------------------------------------------------------
# Tab 8 - User_Flows
# ---------------------------------------------------------------------------
TABS.append(("User_Flows", [
    "Flow_ID", "Flow_name", "Persona", "Step_order", "Route", "Decision_point",
    "Blocker_or_gate", "Current_vs_proposed", "Source_doc",
], [
    ["F1", "Standard happy path", "CS-1 Aditi (salaried)", "1", "/", "-", "-", "Current (~20 screens)", "itr-filing-wireframes/USER_FLOWS.md"],
    ["F1", "Standard happy path", "CS-1 Aditi", "2", "/file/onboarding", "ITR form detection", "-", "Current", "lib/filing/journey.ts"],
    ["F1", "Standard happy path", "CS-1 Aditi", "3", "/file/import", "Form 16 uploaded?", "soft gate", "Current", "FUNNEL_AUDIT_AND_SIMPLIFICATION.md"],
    ["F1", "Standard happy path", "CS-1 Aditi", "4", "/file/regime", "Old vs new", "-", "Current", "lib/filing/journey.ts"],
    ["F1", "Standard happy path", "CS-1 Aditi", "5", "/file/review", "Numbers correct?", "-", "Current", "FUNNEL_AUDIT_AND_SIMPLIFICATION.md"],
    ["F1", "Standard happy path", "CS-1 Aditi", "6", "/file/checkout/plans", "Pay?", "HARD gate (pay after value)", "Current", "FUNNEL_AUDIT_AND_SIMPLIFICATION.md"],
    ["F1", "Standard happy path", "CS-1 Aditi", "7", "/file/companion", "Follow portal steps", "-", "Current", "lib/engine/portalSections.ts"],
    ["F2", "Mismatch block", "CS-3 Meera (AIS mismatch)", "1", "/file/import", "AIS line not in Form 16?", "HARD STOP at mismatch", "Current - the wedge", "itr-filing-wireframes/USER_FLOWS.md"],
    ["F2", "Mismatch block", "CS-3 Meera", "2", "/file/import/mismatch", "Resolve each mismatch", "must resolve before continue", "Current", "lib/filing/reconciliation.ts"],
    ["F2", "Mismatch block", "CS-3 Meera", "3", "/file/review", "Confirm reconciled figures", "-", "Current", "FUNNEL_AUDIT_AND_SIMPLIFICATION.md"],
    ["F3", "Wrong form prevention", "CS-4 Vikram (CG)", "1", "/file/onboarding", "Capital gains present?", "route to ITR-2 (block ITR-1)", "Current", "itr-filing-wireframes/USER_FLOWS.md"],
    ["F3", "Wrong form prevention", "CS-5 Sana (F&O)", "1", "/file/onboarding", "F&O/business present?", "route to ITR-3", "Current", "itr-filing-wireframes/USER_FLOWS.md"],
    ["F4", "Form 16 fast path", "CS-1 Aditi", "1", "/file/import", "Single clean Form 16, no AIS delta", "skip mismatch step", "Current optimization", "FUNNEL_AUDIT_AND_SIMPLIFICATION.md"],
    ["F5", "Proposed 11-screen salaried path", "CS-1 Aditi", "1-11", "Landing -> upload -> instant reconcile -> regime -> review -> pay -> companion", "collapse redundant screens", "merge profiler+import, cut cabrain/everify thin routes", "PROPOSED (target 11-13 vs current ~20)", "FUNNEL_AUDIT_AND_SIMPLIFICATION.md; UX_IMPROVEMENT_PLAN.md"],
    ["F6", "Content -> filing", "Marketing personas", "1", "/learn/[slug] or /blogs/[slug]", "CTA clicked?", "-", "Current", "app/sitemap.ts; lib/content/*"],
    ["F6", "Content -> filing", "Marketing personas", "2", "/file/onboarding", "enter funnel", "-", "Current", "lib/filing/journey.ts"],
    ["F7", "Post-pay companion lifecycle", "All paid", "1", "/file/companion", "Copy each field to portal", "-", "Current - moat", "lib/engine/portalSections.ts"],
    ["F7", "Post-pay companion lifecycle", "All paid", "2", "/file/everify (guide)", "E-verify on portal", "user acts on incometax.gov.in", "Current (thin route)", "itr-filing-wireframes/USER_FLOWS.md"],
]))

# ---------------------------------------------------------------------------
# Tab 9 - Design_References
# ---------------------------------------------------------------------------
TABS.append(("Design_References", [
    "Reference", "Type", "What_to_steal", "What_to_avoid", "File_path", "Figma_or_prototype",
], [
    ["LastMinute ITR Design System", "Design_system", "Tokens, type scale, color, spacing, buttons, cards, tables, forms, empty/error/success, progress, AI assistant components, motion guidelines", "Listed anti-patterns in the doc (e.g., unmotivated motion, density without rhythm)", "docs/DESIGN_SYSTEM.md", "n/a"],
    ["Apple x Quicko north star", "Design_system", "Apple calm + Quicko fintech credibility blend; restraint + clarity", "Over-decoration; hype", "docs/DESIGN_REFERENCE_APPLE_QUICKO.md", "n/a"],
    ["UI Audit", "Design_system", "Catalogued current UI debt + fixes to apply", "Repeating the audited mistakes", "docs/UI_AUDIT.md", "n/a"],
    ["Filing Experience Redesign", "Design_system", "Target redesigned filing flow visual language", "-", "docs/FILING_EXPERIENCE_REDESIGN.md", "n/a"],
    ["shadcn/ui config", "Design_system", "Component primitives + theming base", "Generic look if untokened", "components.json / components/ui/*", "n/a"],
    ["TurboTax", "Competitor_UX", "One question per screen interview; filing-ready % before pay", "Late paywall; refund-max hype", "docs/COMPETITOR_ANALYSIS.md", "n/a"],
    ["Quicko filing journal", "Competitor_UX", "Broker import; expandable income nav; ITD autofill", "Treating mismatch as a minor step", "docs/competitor-audit/QUICKO_FILING_JOURNAL.md", "n/a"],
    ["ClearTax filing journal", "Competitor_UX", "Form 16 visual hero; multi-Form-16 merge", "Tier upsell maze; pay before mismatch", "docs/competitor-audit/CLEARTAX_FILING_JOURNAL.md", "n/a"],
    ["Stripe / Mercury / Ramp", "Global_pattern", "Incremental onboarding; inline doc-parse feedback; contextual trust cards; AI reasoning transparency", "Dev-jargon; non-tax assumptions", "docs/DESIGN_SYSTEM.md (references)", "stripe.com / mercury.com / ramp.com"],
    ["Linear", "Global_pattern", "Purposeful motion; density with rhythm; status affordances", "Motion for its own sake", "docs/DESIGN_SYSTEM.md", "linear.app"],
    ["Wireframe prototype repo", "Wireframe", "Interactive HTML flows (happy / mismatch), mobile + desktop viewports", "-", "itr-filing-wireframes/prototype/index.html", "?flow=happy|mismatch&viewport=mobile|desktop"],
    ["Wireframe USER_FLOWS", "Wireframe", "Mermaid flow diagrams + navigation rules + Figma links", "-", "itr-filing-wireframes/USER_FLOWS.md", "Figma links in doc"],
]))

# ---------------------------------------------------------------------------
# Tab 10 - Competitor_UX_Audits
# ---------------------------------------------------------------------------
TABS.append(("Competitor_UX_Audits", [
    "Competitor", "Audit_type", "Screens_reached", "Auth_walled_at",
    "Key_findings", "Steal_list", "File_path",
], [
    ["Quicko", "Filing_journal", "Landing -> signup -> dashboard -> income heads -> import (partial)", "OTP / login before full import", "Broker autofill is the hook; expandable income nav reduces overwhelm; F&O turnover wizard; mismatch is a step not the moat", "Broker connect strip; expandable income sections; ITD autofill prompt", "docs/competitor-audit/QUICKO_FILING_JOURNAL.md"],
    ["ClearTax", "Filing_journal", "Landing -> Form 16 upload prompt -> tier selection", "Pay/login before mismatch shown", "Form 16 visual hero converts; multi-Form-16 merge; capital gains/crypto coverage; tier upsell maze; pay before value", "Form 16 upload as hero; multi-employer merge UX", "docs/competitor-audit/CLEARTAX_FILING_JOURNAL.md"],
    ["Quicko + ClearTax", "Gap_matrix", "n/a (synthesis)", "n/a", "Side-by-side feature/severity matrix vs LastMinute ITR with Implement/Defer/Keep decisions; our wedge = mismatch-first + honest companion", "Decision column methodology; screen-level feature mapping", "docs/competitor-audit/GAP_MATRIX_2026.md"],
    ["TurboTax + H&R Block", "Content_map", "Marketing + product tours (US, no India filing)", "Account creation for full flow", "Interview-style one-Q-per-screen; persistent help pane; filing-ready % before pay; refund-max hype to avoid", "Interview pacing; help pane pattern; readiness meter", "docs/COMPETITOR_ANALYSIS.md; itr-filing-wireframes/COMPETITOR_GLOBAL.md"],
    ["Landscape (India set)", "CTA_map", "Public marketing pages only", "n/a", "CTA patterns, trust signals, pricing presentation across ClearTax/Quicko/TaxBuddy/Tax2Win/EZTax/myITreturn", "Value-before-pay framing; trust placement at anxiety points", "docs/COMPETITOR_ANALYSIS.md; docs/research/SIMILAR_PLATFORMS.md"],
    ["incometax.gov.in", "Filing_journal", "Full public portal (the real filing surface)", "PAN + OTP login", "Session timeouts, 'Please Wait' loops, confusing utilities - the friction our companion narrates", "Mirror the portal's exact section order in the companion walkthrough", "docs/research/USER_PAINPOINTS_2026.md (T7)"],
]))

# ---------------------------------------------------------------------------
# Tab 11 - Source_Index
# ---------------------------------------------------------------------------
TABS.append(("Source_Index", [
    "File_path", "Title", "Category", "Last_updated", "One_line_summary", "Priority_for_designer",
], [
    ["docs/DESIGN_SYSTEM.md", "Design System", "Design", "2026", "Tokens, components, states, motion - the visual source of truth", "Must_read"],
    ["itr-filing-wireframes/USER_FLOWS.md", "Wireframe User Flows", "IA/Flows", "2026", "Mermaid flows + nav rules + prototype links", "Must_read"],
    ["docs/research/USER_PAINPOINTS_2026.md", "User Pain Points", "Research", "2026", "9 themes synthesized from qna.tax mapped to our routes", "Must_read"],
    ["docs/UX_IMPROVEMENT_PLAN.md", "UX Improvement Plan", "IA/Flows", "2026", "Prioritized UX fixes + proposed simplified funnel", "Must_read"],
    ["docs/FUNNEL_AUDIT_AND_SIMPLIFICATION.md", "Funnel Audit", "IA/Flows", "2026", "Screen-by-screen audit; ~20 -> 11-13 screen proposal", "Must_read"],
    ["docs/COMPETITOR_ANALYSIS.md", "Competitor Analysis", "Competitor", "2026", "ClearTax/Quicko/TurboTax/H&R deep dive + patterns to steal", "Must_read"],
    ["docs/competitor-audit/GAP_MATRIX_2026.md", "Gap Matrix", "Competitor", "2026", "Feature/severity matrix vs Quicko & ClearTax with decisions", "Reference"],
    ["docs/competitor-audit/QUICKO_FILING_JOURNAL.md", "Quicko Filing Journal", "Competitor", "2026", "Step-by-step Quicko journey to the auth wall", "Reference"],
    ["docs/competitor-audit/CLEARTAX_FILING_JOURNAL.md", "ClearTax Filing Journal", "Competitor", "2026", "Step-by-step ClearTax journey to the paywall", "Reference"],
    ["docs/research/SIMILAR_PLATFORMS.md", "Similar Platforms", "Competitor", "2026", "Landscape + premium UI references + borrow/avoid", "Reference"],
    ["docs/case-studies/CASE_STUDIES_2026.md", "Case Studies", "Personas", "2026", "8 taxpayer personas with journeys and must-show screens", "Must_read"],
    ["e2e/fixtures/personas.ts", "Persona Fixtures", "Personas", "2026", "Programmatic persona/draft data for tests + simulations", "Reference"],
    ["lib/copy/competitorInspired.ts", "Competitor-Inspired Copy", "Copy", "2026", "Why-Us pillars, import strip, persona carousel copy", "Reference"],
    ["lib/copy/companion.ts", "Companion Copy", "Copy", "2026", "Companion-safe disclaimers and how-it-works copy (compliance rules)", "Must_read"],
    ["lib/filing/journey.ts", "Journey Steps", "IA/Flows", "2026", "A-F macro steps, process routes, path->step mapping", "Reference"],
    ["app/sitemap.ts", "Sitemap", "IA/Flows", "2026", "Static + dynamic route generation - what is indexed", "Reference"],
    ["components/marketing/FeedbackScreen.tsx", "Feedback Screen", "Feedback", "2026", "In-product 1-5 star + 280-char feedback form spec", "Reference"],
    ["docs/MCKINSEY_ENGAGEMENT/06_ANALYTICS_EVENTS.md", "Analytics Events", "Feedback", "2026", "PostHog companion event catalog (footprint/field events)", "Reference"],
    ["docs/DESIGN_REFERENCE_APPLE_QUICKO.md", "Apple x Quicko Reference", "Design", "2026", "North-star aesthetic blend", "Reference"],
    ["docs/UI_AUDIT.md", "UI Audit", "Design", "2026", "Current UI debt catalogue + fixes", "Reference"],
    ["docs/FILING_EXPERIENCE_REDESIGN.md", "Filing Experience Redesign", "Design", "2026", "Target redesigned filing flow", "Reference"],
    ["docs/brand/BRAND_NAMING_2026.md", "Brand Naming", "Brand", "2026", "Name candidates, collisions, domain analysis", "Archive"],
    ["itr-filing-wireframes/prototype/index.html", "HTML Prototype", "Wireframe", "2026", "Interactive happy/mismatch flow prototype", "Must_read"],
    ["docs/ux-research/_forum_research_raw.json", "Forum Research Raw", "Research", "2026", "Agent-scraped forum quotes (source for Forum_Research tab)", "Reference"],
]))


FORUM_PARTIALS = [
    "_forum_qna.json", "_forum_reddit.json", "_forum_competitors.json",
    "_forum_qna_captured.json", "_forum_reddit_captured.json",
    "_forum_competitors_captured.json",
]


def merge_forum_partials():
    """Combine per-agent forum JSON partials into _forum_research_raw.json."""
    merged = []
    for name in FORUM_PARTIALS:
        path = os.path.join(OUT_DIR, name)
        if not os.path.exists(path):
            continue
        try:
            with open(path, "r", encoding="utf-8") as f:
                part = json.load(f)
            if isinstance(part, list):
                merged.extend(part)
                print(f"  merged {len(part)} rows from {name}")
        except (json.JSONDecodeError, OSError) as exc:
            print(f"  WARNING: could not read {name}: {exc}")
    if merged:
        with open(FORUM_JSON, "w", encoding="utf-8") as f:
            json.dump(merged, f, ensure_ascii=False, indent=2)
        print(f"  wrote {FORUM_JSON} ({len(merged)} combined rows)")


def load_forum_rows():
    """Return (header, rows) for Tab 3 from the agent JSON, deduped vs pain themes."""
    header = [
        "Platform", "Thread_or_post_title", "URL", "Date", "Category",
        "Verbatim_quote", "Pain_summary", "Competitor_mentioned",
        "Relevance_to_us", "Suggested_UI_implication", "Already_in_USER_PAINPOINTS",
    ]
    if not os.path.exists(FORUM_JSON):
        return header, [[
            "(none)", "Forum research JSON not found", "", "", "other",
            "Run the research agents and write docs/ux-research/_forum_research_raw.json",
            "Pending external research", "none", "Low", "Populate before handoff", "No",
        ]]
    with open(FORUM_JSON, "r", encoding="utf-8") as f:
        data = json.load(f)
    # Existing pain theme keywords for dedupe flagging (Tab 2 themes).
    theme_terms = [
        "ais", "form 16", "26as", "reconcil", "mismatch", "regime", "87a",
        "capital gain", "f&o", "intraday", "freelanc", "44ada", "presumptive",
        "nri", "property tds", "portal", "refund", "143", "notice", "which itr",
        "gift", "huf", "pricing",
    ]
    rows = []
    seen = set()
    for item in data:
        url = (item.get("url") or "").strip()
        quote = (item.get("verbatim_quote") or "").strip()
        key = (url, quote[:60].lower())
        if key in seen:
            continue
        seen.add(key)
        blob = " ".join([
            quote.lower(),
            (item.get("pain_summary") or "").lower(),
            (item.get("category") or "").lower(),
            (item.get("thread_title") or "").lower(),
        ])
        already = "Yes" if any(t in blob for t in theme_terms) else "No"
        rows.append([
            item.get("platform", ""),
            item.get("thread_title", ""),
            url,
            item.get("date", "unknown"),
            item.get("category", "other"),
            quote,
            item.get("pain_summary", ""),
            item.get("competitor_mentioned", "none"),
            item.get("relevance_to_us", "Med"),
            item.get("suggested_ui_implication", ""),
            already,
        ])
    return header, rows


def write_csvs():
    os.makedirs(CSV_DIR, exist_ok=True)
    for idx, (name, header, rows) in enumerate(TABS):
        fname = f"{idx:02d}_{name.upper()}.csv"
        path = os.path.join(CSV_DIR, fname)
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(header)
            w.writerows(rows)
        print(f"  wrote {path} ({len(rows)} rows)")


def write_xlsx():
    if not HAS_OPENPYXL:
        print("openpyxl not installed - skipping xlsx; open the CSV bundle instead.")
        return
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    wrap = Alignment(vertical="top", wrap_text=True)
    for name, header, rows in TABS:
        ws = wb.create_sheet(title=name[:31])
        ws.append(header)
        for c in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for row in rows:
            ws.append(row)
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 28
        for c in range(1, len(header) + 1):
            letter = get_column_letter(c)
            longest = max([len(str(header[c - 1]))] + [len(str(r[c - 1])) if c - 1 < len(r) else 0 for r in rows] + [10])
            ws.column_dimensions[letter].width = min(max(longest + 2, 14), 60)
            for r in range(2, len(rows) + 2):
                ws.cell(row=r, column=c).alignment = wrap
    wb.save(XLSX_PATH)
    print(f"  wrote {XLSX_PATH} ({len(TABS)} tabs)")


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    # Combine per-agent forum partials, then load Tab 3 from the merged JSON.
    merge_forum_partials()
    forum_header, forum_rows = load_forum_rows()
    TABS.insert(3, ("Forum_Research", forum_header, forum_rows))
    print(f"Building UX research pack ({len(TABS)} tabs)...")
    write_csvs()
    write_xlsx()
    print("Done.")


if __name__ == "__main__":
    main()
